import json
from openpyxl import load_workbook, Workbook
import os

histDir = r"./controller/history.json"

ctrXlDir = r"./controller/clients.xlsx"

dbXlDir = r"./db/data.xlsx"

cashData = {}


def filterUrls():
    wb = load_workbook(ctrXlDir)

    ws = wb.active

    urls = []

    for row in ws.iter_rows(min_row=1):
        if row[0].value != "ID" and row[0].value is not None:
            if ws["D1"].value == "yes":
                if row[3].value == "yes":
                    urls.append(row[0].value)
            else:
                if row[3].value != "no":
                    urls.append(row[0].value)

    return list(set(urls))


def readUrl(key):
    wb = load_workbook(ctrXlDir)

    ws = wb.active

    if isinstance(key, int) and ws[f"A{key + 1}"].value == key:
        return [ws[f"B{key + 1}"].value, ws[f"C{key + 1}"].value]

    for row in ws.iter_rows(min_row=1):
        if str(row[0].value) == str(key):
            return [row[1].value, row[2].value]


def fetchJobs():
    wb = load_workbook(ctrXlDir)

    ws = wb.active

    return [cell.value.lower() for cell in ws["E"] if cell.value is not None]


def readHistory(key=None):
    with open(histDir, "r") as file:
        try:
            data = json.load(file)

            if key is not None:
                return data.get(f"{key}", [])

            return data
        except:
            return []


def updateHistory(key, val):
    data = readHistory()
    data[f"{key}"] = val

    with open(histDir, "w") as file:
        try:
            json.dump(data, file, indent=4)
        except Exception as e:
            print(e)


def updateDB(key=None, arr=[]):
    global cashData

    if key is None:
        keyList = sorted(list(cashData.keys()))

        if os.path.exists(dbXlDir):
            wb = load_workbook(dbXlDir)
        else:
            wb = Workbook()

        if wb.active:
            ws = wb.active
        else:
            ws = wb.create_sheet()

        if ws["A1"] != "Job Title":
            ws["A1"] = "Job Title"
            ws["B1"] = "Company"
            ws["C1"] = "Location"
            ws["D1"] = "Url"

        for key in keyList:
            for row in cashData[key]:
                ws.append(row)

        wb.save(dbXlDir)

        cashData = {}

        return

    hist = readHistory(key)
    newHist = []
    temp = []

    for item in arr:
        title, _, _, link = item

        newHist.append(link)

        if not link in hist:

            for job in fetchJobs():
                if job in title.lower():
                    temp.append(item)
                    break

    cashData[f"{key}"] = temp

    updateHistory(key, newHist)
